from datetime import timezone

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, ListView, UpdateView
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from ..decorators import student_required
from ..forms import StudentInterestsForm, StudentSignUpForm, TakeQuizForm
from ..models import Quiz, Student, TakenQuiz, User, Room, Message

from django.http import HttpResponse, JsonResponse
from django.views.generic import View
from openpyxl import Workbook
from django.db import connection
from openpyxl.styles import Font


class StudentSignUpView(CreateView):
    model = User
    form_class = StudentSignUpForm
    template_name = 'registration/signup_form.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'student'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return redirect('students:quiz_list')


@method_decorator([login_required, student_required], name='dispatch')
class StudentInterestsView(UpdateView):
    model = Student
    form_class = StudentInterestsForm
    template_name = 'classroom/students/interests_form.html'
    success_url = reverse_lazy('students:quiz_list')

    def get_object(self):
        return self.request.user.student

    def form_valid(self, form):
        messages.success(self.request, 'Interests updated with success!')
        return super().form_valid(form)


@method_decorator([login_required, student_required], name='dispatch')
class QuizListView(ListView):
    model = Quiz
    ordering = ('name',)
    context_object_name = 'quizzes'
    template_name = 'classroom/students/quiz_list.html'

    def get_queryset(self):
        student = self.request.user.student
        student_interests = student.interests.values_list('pk', flat=True)
        taken_quizzes = student.quizzes.values_list('pk', flat=True)
        queryset = Quiz.objects.filter(subject__in=student_interests) \
            .exclude(pk__in=taken_quizzes) \
            .annotate(questions_count=Count('questions')) \
            .filter(questions_count__gt=0)
        return queryset


@method_decorator([login_required, student_required], name='dispatch')
class TakenQuizListView(ListView):
    model = TakenQuiz
    context_object_name = 'taken_quizzes'
    template_name = 'classroom/students/taken_quiz_list.html'

    def get_queryset(self):
        queryset = self.request.user.student.taken_quizzes \
            .select_related('quiz', 'quiz__subject') \
            .order_by('quiz__name')
        return queryset


@login_required
@student_required
def take_quiz(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    student = request.user.student
    team=3

    if student.quizzes.filter(pk=pk).exists():
        return render(request, 'students/taken_quiz.html')

    total_questions = quiz.questions.count()
    unanswered_questions = student.get_unanswered_questions(quiz)
    total_unanswered_questions = unanswered_questions.count()
    progress = 100 - round(((total_unanswered_questions - 1) / total_questions) * 100)
    question = unanswered_questions.first()

    if request.method == 'POST':
        form = TakeQuizForm(question=question, data=request.POST)
        if form.is_valid():
            with transaction.atomic():
                student_answer = form.save(commit=False)
                student_answer.student = student
                student_answer.save()
                if student.get_unanswered_questions(quiz).exists():
                    return redirect('students:take_quiz', pk)
                else:
                    correct_answers = student.quiz_answers.filter(answer__question__quiz=quiz,
                                                                  answer__is_correct=True).count()
                    score = round((correct_answers / total_questions)*10, 2)
                    TakenQuiz.objects.create(student=student, quiz=quiz, score=score)
                    messages.warning(request, 'Прохождение викторины завершено. Ваш результат: %s.' % (
                        score))
                    return redirect('students:quiz_list')
    else:
        form = TakeQuizForm(question=question)

    return render(request, 'classroom/students/take_quiz_form.html', {
        'quiz': quiz,
        'question': question,
        'form': form,
        'progress': progress
    })


@method_decorator([login_required, student_required], name='dispatch')
class ExportToExcelStudentView(View):
    def get(self, request, quiz_id, student_id, *args, **kwargs):
        with connection.cursor() as cursor:
            cursor.execute("select 'Команда: ' || t.name || ' (' ||\
                                   (select group_concat(st.name, ', ')\
                                    from classroom_takenquiz tq1\
                                    inner join classroom_student st on tq1.student_id = st.user_id\
                                    inner join classroom_team t1 on tq1.team_id = t1.id\
                                    where tq1.quiz_id = tq.quiz_id\
                                    and t1.id = tq.team_id\
                                    group by tq1.team_id) || ')' as students,\
                                   'Оценка: ' || tq.score as score\
                            from classroom_takenquiz tq\
                            inner join classroom_team t on tq.team_id = t.id\
                            where tq.quiz_id = %s and tq.student_id = %s\
                            group by tq.team_id", [quiz_id, student_id])
            data = cursor.fetchall()

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Результат викторины"
        ws.column_dimensions[get_column_letter(1)].width = 70
        ws.column_dimensions[get_column_letter(2)].width = 20
        ws.column_dimensions[get_column_letter(3)].width = 18

        for row in data:
            ws.append(row)
        ws.append([])
        ws.append(["Вопрос", "Ответ команды", "Ответ правильный"])
        for row in ws.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.font = Font(bold=True)

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        with connection.cursor() as cursor:
            cursor.execute("select cq.text,\
                                   ca.text,\
                                   case when ca.is_correct = 1 then 'Да' else 'Нет' end as is_correct\
                            from classroom_studentanswer sa\
                            inner join classroom_answer ca on sa.answer_id = ca.id\
                            inner join classroom_question cq on ca.question_id = cq.id\
                            where cq.quiz_id = %s and sa.student_id = %s", [quiz_id, student_id])
            data = cursor.fetchall()
        for row in data:
            ws.append(row)

        for row in ws:
            for cell in row:
                cell.border = border_style

        # Создаем стиль для светло-красного цвета
        fill = PatternFill(start_color="FFFFC0C0", end_color="FFFFC0C0", fill_type="solid")

        # Перебираем ячейки в столбце "Ответ правильный" (колонка C)
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                # Проверяем значение ячейки на "Нет" и применяем стиль, если условие выполняется
                if cell.value == 'Нет':
                    cell.fill = fill

        # Перебираем ячейки и устанавливаем свойство wrap_text для переноса текста
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Student_Result_{request.user.username}.xlsx"'
        wb.save(response)
        return response


@login_required
@student_required
@require_POST
def add_message(request):
    content = request.POST.get('content')
    room_id = 1
    user = request.user

    if content:
        room = Room.objects.get(id=1)
        message = Message.objects.create(room=room, user=user, content=content)
        return JsonResponse({'status': 'ok', 'data': message.serialize()})
    else:
        return JsonResponse({'status': 'error', 'message': 'Content or room ID is missing.'})

