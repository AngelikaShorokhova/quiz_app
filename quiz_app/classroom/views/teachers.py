from datetime import datetime

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Avg, Count
from django.forms import inlineformset_factory
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import (CreateView, DeleteView, DetailView, ListView,
                                  UpdateView)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import PatternFill

from ..decorators import teacher_required
from ..forms import BaseAnswerInlineFormSet, QuestionForm, TeacherSignUpForm, TeamForm
from ..models import Answer, Question, Quiz, User, Team, TeamMembership, Message

from django.http import HttpResponse
from django.views.generic import View
from openpyxl import Workbook
from django.db import connection


class TeacherSignUpView(CreateView):
    model = User
    form_class = TeacherSignUpForm
    template_name = 'registration/signup_form.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'teacher'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return redirect('teachers:quiz_change_list')


@method_decorator([login_required, teacher_required], name='dispatch')
class QuizListView(ListView):
    model = Quiz
    ordering = ('name',)
    context_object_name = 'quizzes'
    template_name = 'classroom/teachers/quiz_change_list.html'

    def get_queryset(self):
        queryset = self.request.user.quizzes \
            .select_related('subject') \
            .annotate(questions_count=Count('questions', distinct=True)) \
            .annotate(taken_count=Count('taken_quizzes', distinct=True))
        return queryset


@method_decorator([login_required, teacher_required], name='dispatch')
class QuizCreateView(CreateView):
    model = Quiz
    fields = ('name', 'subject',)
    template_name = 'classroom/teachers/quiz_add_form.html'

    def form_valid(self, form):
        quiz = form.save(commit=False)
        quiz.owner = self.request.user
        quiz.save()
        return redirect('teachers:quiz_change', quiz.pk)


@method_decorator([login_required, teacher_required], name='dispatch')
class QuizUpdateView(UpdateView):
    model = Quiz
    fields = ('name', 'subject',)
    context_object_name = 'quiz'
    template_name = 'classroom/teachers/quiz_change_form.html'

    def get_context_data(self, **kwargs):
        kwargs['questions'] = self.get_object().questions.annotate(answers_count=Count('answers'))
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        '''
        This method is an implicit object-level permission management
        This view will only match the ids of existing quizzes that belongs
        to the logged in user.
        '''
        return self.request.user.quizzes.all()

    def get_success_url(self):
        return reverse('teachers:quiz_change', kwargs={'pk': self.object.pk})


@method_decorator([login_required, teacher_required], name='dispatch')
class QuizDeleteView(DeleteView):
    model = Quiz
    context_object_name = 'quiz'
    template_name = 'classroom/teachers/quiz_delete_confirm.html'
    success_url = reverse_lazy('teachers:quiz_change_list')

    def delete(self, request, *args, **kwargs):
        quiz = self.get_object()
        return super().delete(request, *args, **kwargs)

    def get_queryset(self):
        return self.request.user.quizzes.all()


@method_decorator([login_required, teacher_required], name='dispatch')
class QuizResultsView(DetailView):
    model = Quiz
    context_object_name = 'quiz'
    template_name = 'classroom/teachers/quiz_results.html'

    def get_context_data(self, **kwargs):
        quiz = self.get_object()
        taken_quizzes = quiz.taken_quizzes.select_related('student__user').order_by('-date')
        total_taken_quizzes = taken_quizzes.count()
        quiz_score = quiz.taken_quizzes.aggregate(average_score=Avg('score'))
        extra_context = {
            'taken_quizzes': taken_quizzes,
            'total_taken_quizzes': total_taken_quizzes,
            'quiz_score': quiz_score
        }
        kwargs.update(extra_context)
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return self.request.user.quizzes.all()


@login_required
@teacher_required
def question_add(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk, owner=request.user)

    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.quiz = quiz
            question.save()
            return redirect('teachers:question_change', quiz.pk, question.pk)
    else:
        form = QuestionForm()

    return render(request, 'classroom/teachers/question_add_form.html', {'quiz': quiz, 'form': form})


@login_required
@teacher_required
def question_change(request, quiz_pk, question_pk):
    quiz = get_object_or_404(Quiz, pk=quiz_pk, owner=request.user)
    question = get_object_or_404(Question, pk=question_pk, quiz=quiz)

    AnswerFormSet = inlineformset_factory(
        Question,  # parent model
        Answer,  # base model
        formset=BaseAnswerInlineFormSet,
        fields=('text', 'is_correct'),
        min_num=2,
        validate_min=True,
        max_num=10,
        validate_max=True
    )

    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        formset = AnswerFormSet(request.POST, instance=question)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.save()
                formset.save()
            return redirect('teachers:quiz_change', quiz.pk)
    else:
        form = QuestionForm(instance=question)
        formset = AnswerFormSet(instance=question)

    return render(request, 'classroom/teachers/question_change_form.html', {
        'quiz': quiz,
        'question': question,
        'form': form,
        'formset': formset
    })


@method_decorator([login_required, teacher_required], name='dispatch')
class QuestionDeleteView(DeleteView):
    model = Question
    context_object_name = 'question'
    template_name = 'classroom/teachers/question_delete_confirm.html'
    pk_url_kwarg = 'question_pk'

    def get_context_data(self, **kwargs):
        question = self.get_object()
        kwargs['quiz'] = question.quiz
        return super().get_context_data(**kwargs)

    def delete(self, request, *args, **kwargs):
        question = self.get_object()
        return super().delete(request, *args, **kwargs)

    def get_queryset(self):
        return Question.objects.filter(quiz__owner=self.request.user)

    def get_success_url(self):
        question = self.get_object()
        return reverse('teachers:quiz_change', kwargs={'pk': question.quiz_id})


@method_decorator([login_required, teacher_required], name='dispatch')
class ExportToExcelView(View):
    def get(self, request, quiz_id, *args, **kwargs):
        with connection.cursor() as cursor:
            cursor.execute("select ct.name as team,\
                                   cs.name as student,\
                                   cq.name as quiz,\
                                   strftime('%%H:%%M %%d.%%m.%%Y',classroom_takenquiz.date) as date,\
                                   round(classroom_takenquiz.score, 2)\
                            from classroom_takenquiz\
                            inner join classroom_student cs on classroom_takenquiz.student_id = cs.user_id\
                            inner join classroom_quiz cq on classroom_takenquiz.quiz_id = cq.id\
                            inner join classroom_user us on cs.user_id = us.id\
                            inner join classroom_team ct on classroom_takenquiz.team_id = ct.id\
                            where cq.id = %s", [quiz_id])
            data = cursor.fetchall()

        # Создаем Excel файл
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Результаты проведения викторины"
        ws1.column_dimensions[get_column_letter(1)].width = 10
        ws1.column_dimensions[get_column_letter(2)].width = 30
        ws1.column_dimensions[get_column_letter(3)].width = 30
        ws1.column_dimensions[get_column_letter(4)].width = 20
        ws1.column_dimensions[get_column_letter(5)].width = 10
        ws1.append(["Команда", "Учащиеся", "Викторина", "Дата прохождения", "Оценка"])
        for row in ws1.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        for row in data:
            ws1.append(row)

        for row in ws1.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                score = cell.value
                if score >= 8:
                    fill = PatternFill(start_color="FFC0FFC0", end_color="FFC0FFC0", fill_type="solid")
                elif score < 4:
                    fill = PatternFill(start_color="FFFFC0C0", end_color="FFFFC0C0", fill_type="solid")
                else:
                    fill = None
                if fill:
                    cell.fill = fill
                cell.border = border_style

        for row in ws1:
            for cell in row:
                cell.border = border_style

        ws2 = wb.create_sheet(title="Статистика по результатам")
        ws2.column_dimensions[get_column_letter(2)].width = 40
        ws2.column_dimensions[get_column_letter(5)].width = 20
        with connection.cursor() as cursor:
            cursor.execute("select t.name,\
                                   (select GROUP_CONCAT(st.name, ', ')\
                                    from classroom_takenquiz tq1\
                                    inner join classroom_student st on tq1.student_id=st.user_id\
                                    inner join classroom_team t1 on tq1.team_id = t1.id\
                                    where tq1.quiz_id = tq.quiz_id\
                                    and t1.id = tq.team_id\
                                    group by tq1.team_id) as students,\
                                   tq.score\
                            from classroom_takenquiz tq\
                            inner join classroom_team t on tq.team_id = t.id\
                            where tq.quiz_id = %s\
                            group by tq.team_id", [quiz_id])
            data = cursor.fetchall()

        # Extract team scores from the data
        team_scores = [(row[0], row[2]) for row in data]

        # Write team scores to the second sheet
        ws2.append(["Команда", "Учащиеся", "Оценка"])
        for row in data:
            ws2.append(row)

        for row in ws2.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.font = Font(bold=True)

        # Create a bar chart based on the team scores
        bar_chart = BarChart()
        values = Reference(ws2, min_col=3, min_row=1, max_row=len(data) + 1)
        categories = Reference(ws2, min_col=1, min_row=2, max_row=len(data) + 2)
        bar_chart.add_data(values, titles_from_data=True)
        bar_chart.set_categories(categories)
        bar_chart.title = "Результаты прохождения"
        bar_chart.x_axis.title = "Команда"
        bar_chart.y_axis.title = "Оценка"
        ws2.add_chart(bar_chart, "A10")

        pie_chart = PieChart()
        labels = ['Отлично', 'Хорошо', 'Удовлетворительно', 'Неудовлетворительно']
        counts = [0, 0, 0, 0]

        for score in team_scores:
            if score[1] >= 8:
                counts[0] += 1
            elif score[1] >= 6:
                counts[1] += 1
            elif score[1] >= 4:
                counts[2] += 1
            else:
                counts[3] += 1

        # Add data to the pie chart
        data = counts
        for i, label in enumerate(labels, start=1):
            ws2[f'E{i}'] = label
            ws2[f'F{i}'] = counts[i - 1]

        for row in ws2:
            for cell in row:
                cell.border = border_style

        total_count = sum(counts)
        percentages = [count / total_count * 100 for count in counts]

        # Add data labels to the pie chart
        data_labels = DataLabelList()
        data_labels.showPercent = True
        pie_chart.dataLabels = data_labels

        # Add data to the pie chart
        values = Reference(ws2, min_col=6, min_row=1, max_row=4)
        categories = Reference(ws2, min_col=5, min_row=1, max_row=4)
        pie_chart.add_data(values, titles_from_data=False)
        pie_chart.set_categories(categories)
        pie_chart.title = "Распределение оценок студентов"

        # Add the pie chart to the third sheet
        ws2.add_chart(pie_chart, "F10")

        response = HttpResponse(content_type='application/ms-excel')
        response[
            'Content-Disposition'] = f'attachment; filename="Quiz_Results_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
        wb.save(response)
        return response


@method_decorator([login_required, teacher_required], name='dispatch')
class TeamListView(View):
    def get(self, request):
        teams = Team.objects.annotate(students_count=Count('students'))
        return render(request, 'classroom/teachers/team_list.html', {'teams': teams})


@method_decorator([login_required, teacher_required], name='dispatch')
class CreateTeamView(View):
    def get(self, request):
        form = TeamForm()
        return render(request, 'classroom/teachers/team_add_form.html', {'form': form})

    def post(self, request):
        form = TeamForm(request.POST)
        if form.is_valid():
            team = form.save(commit=False)
            team.save()
            form.save_m2m()
            return redirect('teachers:team_list')
        return render(request, 'classroom/teachers/team_add_form.html', {'form': form})


def team_view(request, team_id):
    user_messages = Message.objects.filter(room_id=1)
    team = Team.objects.get(id=team_id)
    context = {
        'team': team,
        'user_messages': user_messages
    }
    return render(request, 'classroom/teachers/team_detail.html', context)
